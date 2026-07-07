"""
Работа с Google Sheets для табеля учёта рабочего времени.

Логика:
  - 08:00 — заполнить все ячейки текущего дня кодом "Я" (явка)
  - в течение дня — бригадир меняет ячейки через выпадающий список
  - 20:00 — если за день НИ ОДНА ячейка не изменилась (у всех осталось "Я"),
            отправить вопрос; при отсутствии подтверждения — обнулить день.
"""

import calendar
import json
import os
import time
from datetime import datetime

import gspread
from google.oauth2.service_account import Credentials

# --- Настройки ---
SPREADSHEET_ID = "1d7YqIAqWL9_cQQ7JpxqD_qV69q1NpVO3u58BzDlK73M"

# JSON-ключ service account.
# На Railway кладётся в переменную окружения GOOGLE_CREDENTIALS (весь JSON).
# Локально можно положить файл service_account.json рядом с кодом.
CREDENTIALS_FILE = "service_account.json"

# Коды статусов (СТАРАЯ модель — пока оставлены для совместимости этапа перехода)
CODE_PRESENT = "Я"   # явка
CODE_ABSENT = "Н"    # неявка
CODE_SICK = "Б"      # больничный
CODE_VACATION = "О"  # отпуск
CODE_WEEKEND = "В"   # выходной
ALL_CODES = [CODE_PRESENT, CODE_ABSENT, CODE_SICK, CODE_VACATION, CODE_WEEKEND]

# --- НОВАЯ модель ДЕНЬ/НОЧЬ ---
# Дневной слот
DN_DAY = "Д"       # работал день
DN_REST = "О"      # отдых
DN_SICK = "Б"      # больничный
DN_ROTATION = "МЖ" # межвахта
DN_ABSENT = "Н"    # неявка
DN_MIGR = "МУ"     # миграционный учёт
DN_WEEKEND = "В"   # плановый выходной
# Ночной слот
DN_NIGHT = "НЧ"    # работал ночь

DAY_CODES = [DN_DAY, DN_REST, DN_SICK, DN_ROTATION, DN_ABSENT, DN_MIGR, DN_WEEKEND]
NIGHT_CODES = [DN_NIGHT, DN_REST]
# Причины отсутствия (для шага «оставшиеся»)
REASON_CODES = [DN_ABSENT, DN_SICK, DN_ROTATION, DN_MIGR, DN_WEEKEND]

# Русские названия месяцев = названия листов
MONTHS_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]

# Структура листа (модель ДЕНЬ/НОЧЬ):
#   строка 1 — заголовок месяца
#   строка 2 — числа дней (объединены над парой Д|Н)
#   строка 3 — подписи слотов: Д | Н | Д | Н ...
#   строки 4+ — сотрудники: A=№, B=ФИО, далее пары день/ночь
FIRST_DATA_ROW = 4
NUM_COL = 1           # столбец A = №
NAME_COL = 2          # столбец B = ФИО
FIRST_DAY_COL = 3     # столбец C = день 1 (дневной слот)

# Лист-справочник сотрудников
EMP_SHEET = "Сотрудники"
EMP_STATUS_ACTIVE = "активен"
EMP_STATUS_FIRED = "уволен"

# Лист пользователей бота (доступ)
USERS_SHEET = "Пользователи"
ROLE_ADMIN = "админ"
ROLE_FOREMAN = "прораб"

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


class SheetsBusyError(Exception):
    """Google Sheets API временно недоступен (превышена квота запросов)."""
    pass


def _is_quota_error(e: Exception) -> bool:
    return isinstance(e, gspread.exceptions.APIError) and "429" in str(e)


def _credentials():
    """
    Загружает Credentials.
    Приоритет: переменная окружения GOOGLE_CREDENTIALS (для Railway),
    иначе — локальный файл service_account.json.
    """
    raw = os.getenv("GOOGLE_CREDENTIALS")
    if raw:
        info = json.loads(raw)
        return Credentials.from_service_account_info(info, scopes=SCOPES)
    return Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)


_cached_client = None
_cached_spreadsheet = None


def _client():
    global _cached_client
    if _cached_client is None:
        _cached_client = gspread.authorize(_credentials())
    return _cached_client


def _open():
    global _cached_spreadsheet
    if _cached_spreadsheet is None:
        _cached_spreadsheet = _client().open_by_key(SPREADSHEET_ID)
    return _cached_spreadsheet


_ws_cache = {}


def _worksheet_for(date: datetime):
    """Лист месяца (кэшируется по названию, чтобы не читать метаданные книги)."""
    title = MONTHS_RU[date.month - 1]
    if title not in _ws_cache:
        _ws_cache[title] = _open().worksheet(title)
    return _ws_cache[title]


def _day_col(date: datetime) -> int:
    """Столбец ДНЕВНОГО слота для числа (1-based). Пары: С=1Д, D=1Н, E=2Д..."""
    return FIRST_DAY_COL + (date.day - 1) * 2


def _night_col(date: datetime) -> int:
    """Столбец НОЧНОГО слота для числа (1-based)."""
    return _day_col(date) + 1


# Совместимость со старым именем (если где-то ещё вызывается)
def _day_column(date: datetime) -> int:
    return _day_col(date)


def _employee_count(ws) -> int:
    """Сколько строк-сотрудников на листе."""
    names = ws.col_values(NAME_COL)  # включая шапку
    # names[0]=заголовок месяца (стр.1), names[1]=шапка (стр.2), дальше ФИО
    return max(0, len(names) - (FIRST_DATA_ROW - 1))


_grid_cache = {"data": None, "ts": 0, "sheet": None}
_GRID_TTL = 15  # секунд


def _read_grid(date: datetime | None = None):
    """
    Читает весь лист месяца (кэш 15 сек). Возвращает (ws, grid).
    При отметках кэш обновляется локально через _grid_set.
    """
    date = date or datetime.now()
    ws = _worksheet_for(date)
    now = time.time()
    if (_grid_cache["sheet"] != ws.title
            or now - _grid_cache["ts"] > _GRID_TTL
            or _grid_cache["data"] is None):
        _grid_cache["data"] = ws.get_all_values()
        _grid_cache["ts"] = now
        _grid_cache["sheet"] = ws.title
    return ws, _grid_cache["data"]


def _grid_set(ws, row: int, col: int, value: str):
    """Локально обновляет кэш grid после записи ячейки (row/col 1-based)."""
    if _grid_cache["sheet"] != ws.title or _grid_cache["data"] is None:
        return
    grid = _grid_cache["data"]
    ri, ci = row - 1, col - 1
    while len(grid) <= ri:
        grid.append([])
    while len(grid[ri]) <= ci:
        grid[ri].append("")
    grid[ri][ci] = value


def check_rotation_return_conflict(name: str, date: datetime | None = None) -> str | None:
    """
    Проверка перед простановкой ДНЯ (Д): если вчера у сотрудника стояла
    межвахта (МЖ), значит возврат происходит без штатного оформления.
    Возвращает строгий текст предупреждения или None.
    """
    date = date or datetime.now()
    from datetime import timedelta
    yday = date - timedelta(days=1)
    try:
        _, ygrid = _read_grid(yday)
        yd_idx = _day_col(yday) - 1
        for r in ygrid[FIRST_DATA_ROW - 1:]:
            if len(r) > NAME_COL - 1 and r[NAME_COL - 1].strip() == name.strip():
                yval = r[yd_idx].strip() if len(r) > yd_idx else ""
                if yval == DN_ROTATION:
                    return (
                        f"ВНИМАНИЕ: {name} был на межвахте. Возврат сотрудника "
                        f"на объект без надлежащего уведомления и постановки на "
                        f"миграционный учёт является нарушением миграционного "
                        f"законодательства РФ.")
                break
    except Exception:
        pass
    return None


def check_day_conflict(name: str, date: datetime | None = None) -> str | None:
    """
    Проверки перед простановкой ДНЯ (Д) сотруднику:
      - если у него в этот день уже стоит ночь (НЧ) → конфликт день/ночь
      - если вчера была ночь (НЧ) → работа сразу после ночи
    Возвращает текст предупреждения или None, если всё чисто.
    """
    date = date or datetime.now()
    ws, grid = _read_grid(date)
    n_idx = _night_col(date) - 1
    # текущий ночной слот
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) > NAME_COL - 1 and r[NAME_COL - 1].strip() == name.strip():
            nval = r[n_idx].strip() if len(r) > n_idx else ""
            if nval == DN_NIGHT:
                return f"{name} уже отмечен в НОЧЬ за этот день."
            break
    # вчерашняя ночь
    from datetime import timedelta
    yday = date - timedelta(days=1)
    try:
        _, ygrid = _read_grid(yday)
        yn_idx = _night_col(yday) - 1
        for r in ygrid[FIRST_DATA_ROW - 1:]:
            if len(r) > NAME_COL - 1 and r[NAME_COL - 1].strip() == name.strip():
                yval = r[yn_idx].strip() if len(r) > yn_idx else ""
                if yval == DN_NIGHT:
                    return f"{name} вчера работал в НОЧЬ, положен отдых."
                break
    except Exception:
        pass
    return None


def check_night_conflict(name: str, date: datetime | None = None) -> str | None:
    """
    Проверка перед простановкой НОЧИ (НЧ):
      - если сотрудник уже отработал день (Д) → конфликт день/ночь.
    Возвращает текст предупреждения или None.
    """
    date = date or datetime.now()
    ws, grid = _read_grid(date)
    d_idx = _day_col(date) - 1
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) > NAME_COL - 1 and r[NAME_COL - 1].strip() == name.strip():
            dval = r[d_idx].strip() if len(r) > d_idx else ""
            if dval == DN_DAY:
                return f"{name} уже отработал ДЕНЬ за эту дату."
            break
    return None


def mark_day(name: str, date: datetime | None = None) -> bool:
    """Прораб отметил присутствующего днём: ДЕНЬ=Д."""
    date = date or datetime.now()
    ws = _worksheet_for(date)
    row = _row_by_name(ws, name)
    if row is None:
        return False
    col = _day_col(date)
    ws.update_cell(row, col, DN_DAY)
    _grid_set(ws, row, col, DN_DAY)
    return True


def mark_night(name: str, date: datetime | None = None) -> bool:
    """Прораб отметил ночную смену: НОЧЬ=НЧ, ДЕНЬ=О (днём отдыхал)."""
    date = date or datetime.now()
    ws = _worksheet_for(date)
    row = _row_by_name(ws, name)
    if row is None:
        return False
    dcol, ncol = _day_col(date), _night_col(date)
    ws.update_cell(row, dcol, DN_REST)
    ws.update_cell(row, ncol, DN_NIGHT)
    _grid_set(ws, row, dcol, DN_REST)
    _grid_set(ws, row, ncol, DN_NIGHT)
    return True


def set_reason(name: str, code: str, date: datetime | None = None) -> bool:
    """Причина отсутствия в дневной слот: Н / Б / МЖ."""
    date = date or datetime.now()
    ws = _worksheet_for(date)
    row = _row_by_name(ws, name)
    if row is None:
        return False
    col = _day_col(date)
    ws.update_cell(row, col, code)
    _grid_set(ws, row, col, code)
    return True


def set_rest(name: str, date: datetime | None = None) -> bool:
    """Автоотдых с ночи: ДЕНЬ=О."""
    date = date or datetime.now()
    ws = _worksheet_for(date)
    row = _row_by_name(ws, name)
    if row is None:
        return False
    col = _day_col(date)
    ws.update_cell(row, col, DN_REST)
    _grid_set(ws, row, col, DN_REST)
    return True


def clear_day_slot(name: str, date: datetime | None = None) -> bool:
    """Очистка дневного слота сотрудника (Д→пусто). Ночной не трогаем."""
    date = date or datetime.now()
    ws = _worksheet_for(date)
    row = _row_by_name(ws, name)
    if row is None:
        return False
    col = _day_col(date)
    ws.update_cell(row, col, "")
    _grid_set(ws, row, col, "")
    return True


def clear_night_slot(name: str, date: datetime | None = None) -> bool:
    """Очистка ночного слота сотрудника (Н→пусто). Дневной не трогаем."""
    date = date or datetime.now()
    ws = _worksheet_for(date)
    row = _row_by_name(ws, name)
    if row is None:
        return False
    col = _night_col(date)
    ws.update_cell(row, col, "")
    _grid_set(ws, row, col, "")
    return True


def get_marked_day(date: datetime | None = None) -> list[str]:
    """Активные, у кого дневной слот НЕ пуст (для очистки в «Утро»)."""
    date = date or datetime.now()
    ws, grid = _read_grid(date)
    col_idx = _day_col(date) - 1
    active = get_employees(date)
    val = {}
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) > NAME_COL - 1:
            nm = r[NAME_COL - 1].strip()
            val[nm] = r[col_idx].strip() if len(r) > col_idx else ""
    return [n for n in active if val.get(n, "")]


def get_marked_night(date: datetime | None = None) -> list[str]:
    """Активные, у кого ночной слот = НЧ (для очистки в «Вечер»)."""
    date = date or datetime.now()
    ws, grid = _read_grid(date)
    col_idx = _night_col(date) - 1
    active = get_employees(date)
    val = {}
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) > NAME_COL - 1:
            nm = r[NAME_COL - 1].strip()
            val[nm] = r[col_idx].strip() if len(r) > col_idx else ""
    return [n for n in active if val.get(n, "") == DN_NIGHT]


def morning_progress(date: datetime | None = None) -> dict:
    """
    Состояние утренней отметки:
      marked  — сколько с непустым дневным слотом
      unmarked — сколько с пустым
    Прерванная отметка = marked > 0 И unmarked > 0.
    """
    date = date or datetime.now()
    ws, grid = _read_grid(date)
    col_idx = _day_col(date) - 1
    active = set(get_employees(date))
    marked = unmarked = 0
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) <= NAME_COL - 1:
            continue
        nm = r[NAME_COL - 1].strip()
        if nm not in active:
            continue
        v = r[col_idx].strip() if len(r) > col_idx else ""
        if v:
            marked += 1
        else:
            unmarked += 1
    return {"marked": marked, "unmarked": unmarked,
            "interrupted": marked > 0 and unmarked > 0}


def get_night_rest(date: datetime | None = None) -> list[str]:
    """
    Кто вчера работал в ночь (НОЧЬ=НЧ) — тем сегодня положен отдых днём.
    Корректно смотрит в прошлый месяц при переходе через 1-е число.
    """
    date = date or datetime.now()
    from datetime import timedelta
    yday = date - timedelta(days=1)
    try:
        ws, grid = _read_grid(yday)
    except Exception:
        return []
    night_idx = _night_col(yday) - 1  # 0-based
    active = set(get_employees(date))
    result = []
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) <= NAME_COL - 1:
            continue
        name = r[NAME_COL - 1].strip()
        if name in active and len(r) > night_idx and r[night_idx].strip() == DN_NIGHT:
            result.append(name)
    return result


def get_day_slot(name: str, date: datetime | None = None) -> str:
    """Текущее значение дневного слота сотрудника."""
    date = date or datetime.now()
    ws, grid = _read_grid(date)
    col_idx = _day_col(date) - 1
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) > NAME_COL - 1 and r[NAME_COL - 1].strip() == name.strip():
            return r[col_idx].strip() if len(r) > col_idx else ""
    return ""


def get_unmarked_day(date: datetime | None = None) -> list[str]:
    """Активные, у кого дневной слот ПУСТ (ещё не отмечены утром)."""
    date = date or datetime.now()
    ws, grid = _read_grid(date)
    col_idx = _day_col(date) - 1
    active = get_employees(date)
    present = {}
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) > NAME_COL - 1:
            nm = r[NAME_COL - 1].strip()
            present[nm] = r[col_idx].strip() if len(r) > col_idx else ""
    return [n for n in active if not present.get(n, "")]


def clear_all_day(date: datetime | None = None) -> int:
    """
    ТЕСТОВАЯ: очищает ОБА слота (день+ночь) у всех активных за день.
    Батч-запись. Возвращает число очищенных сотрудников.
    """
    date = date or datetime.now()
    active = get_employees(date)
    ws = _worksheet_for(date)
    dcol, ncol = _day_col(date), _night_col(date)
    cells = []
    for name in active:
        row = _row_by_name(ws, name)
        if row:
            cells.append(gspread.Cell(row, dcol, ""))
            cells.append(gspread.Cell(row, ncol, ""))
            _grid_set(ws, row, dcol, "")
            _grid_set(ws, row, ncol, "")
    if cells:
        ws.update_cells(cells)
    return len(active)


def fill_unmarked_absent(date: datetime | None = None) -> int:
    """
    Всем активным с пустым дневным слотом ставит Н (неявка).
    Батч-запись одним запросом. Возвращает число проставленных.
    """
    date = date or datetime.now()
    unmarked = get_unmarked_day(date)
    if not unmarked:
        return 0
    ws = _worksheet_for(date)
    col = _day_col(date)
    cells = []
    for name in unmarked:
        row = _row_by_name(ws, name)
        if row:
            cells.append(gspread.Cell(row, col, DN_ABSENT))
            _grid_set(ws, row, col, DN_ABSENT)
    if cells:
        ws.update_cells(cells)
    return len(cells)


def get_not_worked_day(date: datetime | None = None) -> list[str]:
    """
    Для вечера: активные, кто НЕ работал днём (слот != Д).
    Их можно поставить в ночь.
    """
    date = date or datetime.now()
    ws, grid = _read_grid(date)
    col_idx = _day_col(date) - 1
    active = get_employees(date)
    day_val = {}
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) > NAME_COL - 1:
            nm = r[NAME_COL - 1].strip()
            day_val[nm] = r[col_idx].strip() if len(r) > col_idx else ""
    return [n for n in active if day_val.get(n, "") != DN_DAY]


def _all_month_names(date: datetime | None = None) -> list[str]:
    """Все ФИО из листа месяца (по порядку строк)."""
    date = date or datetime.now()
    ws = _worksheet_for(date)
    names = ws.col_values(NAME_COL)
    return names[FIRST_DATA_ROW - 1:]


_status_cache = {"data": None, "ts": 0}
_STATUS_TTL = 30  # секунд


def get_status_list(force: bool = False) -> list[dict]:
    """
    Читает лист «Сотрудники» (с кэшем на 30 сек).
    Возвращает список {"name", "status", "fired_date"} по порядку.
    Если листа нет — пустой список.
    """
    now = time.time()
    if not force and _status_cache["data"] is not None \
            and now - _status_cache["ts"] < _STATUS_TTL:
        return _status_cache["data"]
    try:
        ws = _open().worksheet(EMP_SHEET)
    except Exception:
        return []
    rows = ws.get_all_values()[1:]  # без шапки
    result = []
    for r in rows:
        if len(r) >= 2 and r[1].strip():
            result.append({
                "name": r[1].strip(),
                "status": (r[2].strip() if len(r) > 2 else EMP_STATUS_ACTIVE),
                "fired_date": (r[3].strip() if len(r) > 3 else ""),
            })
    _status_cache["data"] = result
    _status_cache["ts"] = now
    return result


def get_employees(date: datetime | None = None) -> list[str]:
    """
    Список ФИО активных сотрудников.
    Если есть лист «Сотрудники» — берём только активных оттуда.
    Иначе — все из листа месяца (обратная совместимость).
    """
    status = get_status_list()
    if status:
        active = [e["name"] for e in status if e["status"] == EMP_STATUS_ACTIVE]
        return sorted(active, key=lambda n: n.strip().lower())
    return _all_month_names(date)


def get_fired() -> list[dict]:
    """Список уволенных: [{"name", "fired_date"}, ...]."""
    return [
        {"name": e["name"], "fired_date": e["fired_date"]}
        for e in get_status_list()
        if e["status"] == EMP_STATUS_FIRED
    ]


def employee_exists(name: str) -> bool:
    """Есть ли уже такой сотрудник (активный или уволенный) в «Сотрудники»."""
    nm = name.strip().lower()
    return any(e["name"].strip().lower() == nm for e in get_status_list())


def _name_key(name: str) -> str:
    """Фамилия+Имя (первые 2 токена) в нижнем регистре — для fuzzy-сравнения."""
    parts = name.split()
    return " ".join(parts[:2]).strip().lower() if len(parts) >= 2 else name.strip().lower()


def find_fuzzy_matches(name: str, status_list: list[dict] | None = None) -> list[dict]:
    """Существующие сотрудники с тем же Фамилия+Имя, но другим полным ФИО."""
    if status_list is None:
        status_list = get_status_list()
    key = _name_key(name)
    nm_full = name.strip().lower()
    return [e for e in status_list
            if _name_key(e["name"]) == key and e["name"].strip().lower() != nm_full]


_sheet_meta_cache = {"sheet_ids": None, "ts": 0}
_SHEET_META_TTL = 3600  # структура листов не меняется на лету, кэшируем на час


def _get_sheet_ids(service) -> dict:
    """sheetId по названию листа, кэш на час — метаданные не меняются между
    добавлениями сотрудников, незачем читать их заново на каждого."""
    now = time.time()
    if _sheet_meta_cache["sheet_ids"] is None or now - _sheet_meta_cache["ts"] > _SHEET_META_TTL:
        meta = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        _sheet_meta_cache["sheet_ids"] = {
            s["properties"]["title"]: s["properties"]["sheetId"] for s in meta["sheets"]
        }
        _sheet_meta_cache["ts"] = now
    return _sheet_meta_cache["sheet_ids"]


def _fetch_last_rows(service) -> dict:
    """
    Одним batchGet-запросом получает номер последней занятой строки (по ФИО)
    для всех 12 листов месяцев разом — вместо 12 отдельных get_all_values().
    Возвращает {название_месяца: последняя_строка}.
    """
    ranges = [f"{m}!B{FIRST_DATA_ROW}:B" for m in MONTHS_RU]
    resp = service.spreadsheets().values().batchGet(
        spreadsheetId=SPREADSHEET_ID, ranges=ranges).execute()
    result = {}
    for month, vr in zip(MONTHS_RU, resp.get("valueRanges", [])):
        values = vr.get("values", [])
        last = FIRST_DATA_ROW - 1
        for i, row in enumerate(values):
            if row and row[0].strip():
                last = FIRST_DATA_ROW + i
        result[month] = last
    return result


def _ws_by_title(sp, title: str):
    """Worksheet по названию с кэшем на весь процесс (см. _ws_cache)."""
    if title not in _ws_cache:
        _ws_cache[title] = sp.worksheet(title)
    return _ws_cache[title]


def add_employee(name: str, year: int = 2026, _skip_exists_check: bool = False,
                  _next_num: int | None = None, _emp_new_row: int | None = None,
                  _last_rows: dict | None = None) -> bool:
    """
    Добавляет нового сотрудника:
      - в лист «Сотрудники» (в конец, статус активен)
      - строкой в конец всех 12 листов месяцев (№, ФИО)
      - настраивает выпадающие списки Д/Н на новую строку
    Возвращает False, если уже существует.

    _skip_exists_check: пропустить employee_exists() (и лишний API-запрос),
    когда вызывающий код уже сам проверил уникальность имени по локальным
    данным (см. add_employees_from_xlsx).

    _next_num / _emp_new_row / _last_rows: предвычисленное состояние для
    пакетной загрузки — если передано, add_employee НЕ читает «Сотрудники»
    и не делает _fetch_last_rows() заново, а использует переданные значения.
    Вызывающий код (add_employees_from_xlsx) обязан сам инкрементировать
    эти значения после каждого успешного добавления — иначе следующий
    сотрудник попадёт в ту же строку.
    """
    import calendar as _cal
    from googleapiclient.discovery import build as _build

    name = " ".join(name.split())  # нормализуем пробелы
    if not _skip_exists_check and employee_exists(name):
        return False

    sp = _open()

    # 1. Лист «Сотрудники» — номер новой строки (из кэша пачки либо читаем)
    if _next_num is None or _emp_new_row is None:
        try:
            sp.worksheet(EMP_SHEET)
        except Exception:
            return False
        emp_rows = sp.worksheet(EMP_SHEET).get_all_values()
        next_num = len([r for r in emp_rows[1:] if r and r[0].strip()]) + 1
        emp_new_row = len(emp_rows) + 1
    else:
        next_num = _next_num
        emp_new_row = _emp_new_row
    hire_date = datetime.now().strftime("%d.%m.%Y")
    _status_cache["data"] = None

    # 2. Собираем ВСЕ записи значений (Сотрудники + 12 месяцев) в ОДИН
    # запрос values().batchUpdate — вместо 13 отдельных write-вызовов.
    service = _build("sheets", "v4", credentials=_credentials())
    sheet_ids = _get_sheet_ids(service)
    last_rows = _last_rows if _last_rows is not None else _fetch_last_rows(service)

    # A=№ B=ФИО C=статус D=увольнение E=межвахта F=дата приёма
    value_data = [{
        "range": f"{EMP_SHEET}!A{emp_new_row}:F{emp_new_row}",
        "values": [[str(next_num), name, EMP_STATUS_ACTIVE, "", "", hire_date]],
    }]

    requests = []
    for month_idx, month_name in enumerate(MONTHS_RU, 1):
        try:
            ws_m = _ws_by_title(sp, month_name)
        except Exception:
            continue
        new_row = last_rows.get(month_name, FIRST_DATA_ROW - 1) + 1
        value_data.append({
            "range": f"{month_name}!A{new_row}:B{new_row}",
            "values": [[next_num, name]],
        })

        days = _cal.monthrange(year, month_idx)[1]
        sheet_id = sheet_ids.get(month_name)
        if sheet_id is None:
            continue
        pink = {"red": 0.99, "green": 0.89, "blue": 0.84}
        thick = {"style": "SOLID_THICK", "color": {"red": 0, "green": 0, "blue": 0}}
        thin = {"style": "SOLID", "color": {"red": 0.6, "green": 0.6, "blue": 0.6}}
        black_thin = {"style": "SOLID", "color": {"red": 0, "green": 0, "blue": 0}}
        # Граница на № и ФИО — раньше её тут не было вообще, хотя в
        # исходном шаблоне (строки 1-37) она есть на каждой строке.
        requests.append({
            "updateBorders": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": new_row - 1, "endRowIndex": new_row,
                    "startColumnIndex": 0, "endColumnIndex": 2,
                },
                "top": black_thin, "bottom": black_thin,
                "left": black_thin, "right": black_thin,
                "innerVertical": black_thin,
            }
        })
        for d in range(days):
            day_col = (FIRST_DAY_COL - 1) + d * 2  # 0-based
            night_col = day_col + 1
            requests.append(_dv_row(sheet_id, new_row - 1, day_col, DAY_CODES))
            requests.append(_dv_row(sheet_id, new_row - 1, night_col, NIGHT_CODES))
            # розовый фон выходных для новой строки
            if _cal.weekday(year, month_idx, d + 1) >= 5:
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": new_row - 1, "endRowIndex": new_row,
                            "startColumnIndex": day_col, "endColumnIndex": day_col + 2,
                        },
                        "cell": {"userEnteredFormat": {"backgroundColor": pink}},
                        "fields": "userEnteredFormat.backgroundColor",
                    }
                })
            # границы пары для новой строки
            requests.append({
                "updateBorders": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": new_row - 1, "endRowIndex": new_row,
                        "startColumnIndex": day_col, "endColumnIndex": day_col + 2,
                    },
                    "left": thick, "right": thick, "bottom": thin,
                    "innerVertical": thin,
                }
            })

    # Одним вызовом пишем ВСЕ значения (Сотрудники + 12 месяцев)
    service.spreadsheets().values().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={"valueInputOption": "USER_ENTERED", "data": value_data}
    ).execute()

    if requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID, body={"requests": requests}).execute()

    # сбрасываем кэши данных (не _ws_cache — сами объекты Worksheet не
    # устаревают, а их сброс на каждом добавлении сводит на нет кэш из
    # _ws_by_title при массовой загрузке нескольких сотрудников подряд)
    _grid_cache["data"] = None
    _rowmap_cache["data"] = {}
    return True


def _dv_row(sheet_id, row_idx0, col_idx0, codes):
    """Data validation для одной ячейки (row/col 0-based)."""
    return {
        "setDataValidation": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": row_idx0, "endRowIndex": row_idx0 + 1,
                "startColumnIndex": col_idx0, "endColumnIndex": col_idx0 + 1,
            },
            "rule": {
                "condition": {
                    "type": "ONE_OF_LIST",
                    "values": [{"userEnteredValue": c} for c in codes],
                },
                "showCustomUi": True, "strict": False,
            },
        }
    }


_rowmap_cache = {"data": {}, "ts": 0, "sheet": None}
_ROWMAP_TTL = 60  # секунд


def _row_by_name(ws, name: str) -> int | None:
    """
    Номер строки сотрудника по ФИО. Карта ФИО→строка кэшируется на 60 сек
    для листа, чтобы не читать столбец на каждый тап.
    """
    now = time.time()
    if (_rowmap_cache["sheet"] != ws.title
            or now - _rowmap_cache["ts"] > _ROWMAP_TTL
            or not _rowmap_cache["data"]):
        names = ws.col_values(NAME_COL)
        m = {}
        for i, n in enumerate(names):
            if i >= FIRST_DATA_ROW - 1 and n.strip():
                m[n.strip()] = i + 1
        _rowmap_cache["data"] = m
        _rowmap_cache["ts"] = now
        _rowmap_cache["sheet"] = ws.title
    return _rowmap_cache["data"].get(name.strip())


def set_rotation_return(name: str, return_date: str) -> bool:
    """Сохраняет дату возврата с межвахты в лист «Сотрудники» (столбец E)."""
    try:
        ws = _open().worksheet(EMP_SHEET)
    except Exception:
        return False
    grid = ws.get_all_values()
    for i, r in enumerate(grid):
        if i == 0:
            continue
        if len(r) >= 2 and r[1].strip() == name.strip():
            ws.update_cell(i + 1, 5, return_date)  # E = Межвахта до
            _status_cache["data"] = None
            return True
    return False


def get_rotation_reminders(days_before: int = 3) -> list[dict]:
    """
    Возвращает тех, кто возвращается с межвахты в пределах days_before дней.
    [{"name", "return_date"}]. Дата в формате ДД.ММ.
    """
    from datetime import timedelta
    try:
        ws = _open().worksheet(EMP_SHEET)
    except Exception:
        return []
    grid = ws.get_all_values()
    today = datetime.now().date()
    result = []
    for i, r in enumerate(grid):
        if i == 0 or len(r) < 5:
            continue
        raw = r[4].strip()
        if not raw:
            continue
        # парсим ДД.ММ или ДД.ММ.ГГГГ
        parts = raw.split(".")
        try:
            d = int(parts[0]); m = int(parts[1])
            y = int(parts[2]) if len(parts) > 2 else today.year
            ret = datetime(y, m, d).date()
        except Exception:
            continue
        delta = (ret - today).days
        if 0 <= delta <= days_before:
            result.append({"name": r[1].strip(), "return_date": raw})
    return result


def get_current_status(emp_index: int, date: datetime | None = None) -> str:
    """
    Возвращает текущее значение ДНЕВНОГО слота сотрудника.
    Нужно для предупреждения о перезаписи.
    """
    date = date or datetime.now()
    active = get_employees(date)
    if emp_index >= len(active):
        return ""
    name = active[emp_index]
    ws = _worksheet_for(date)
    row = _row_by_name(ws, name)
    if row is None:
        return ""
    col = _day_col(date)
    val = ws.cell(row, col).value
    return (val or "").strip()


def set_status(emp_index: int, code: str, date: datetime | None = None):
    """
    Ставит статус сотруднику за конкретный день.
    emp_index — индекс в списке get_employees() (активные).
    Запись идёт по ФИО (поиск строки в листе месяца), чтобы уволенные
    не сдвигали адресацию.
    Возвращает (ФИО, код).
    """
    date = date or datetime.now()
    active = get_employees(date)
    if emp_index >= len(active):
        return None, code
    name = active[emp_index]

    ws = _worksheet_for(date)
    row = _row_by_name(ws, name)
    if row is None:
        return name, code
    col = _day_column(date)
    ws.update_cell(row, col, code)
    return name, code


def day_summary(date: datetime | None = None) -> dict:
    """
    Сводка за день (модель день/ночь) по активным сотрудникам.
    Считает дневной и ночной слоты. Один запрос на лист.
    """
    date = date or datetime.now()
    active = set(get_employees(date))
    ws = _worksheet_for(date)
    grid = ws.get_all_values()

    d_idx = _day_col(date) - 1
    n_idx = _night_col(date) - 1

    day_work = night_work = rest = sick = rotation = absent = migr = 0
    absent_list = []

    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) < NAME_COL:
            continue
        name = r[NAME_COL - 1].strip()
        if not name or name not in active:
            continue
        dval = r[d_idx].strip() if len(r) > d_idx else ""
        nval = r[n_idx].strip() if len(r) > n_idx else ""

        if dval == DN_DAY:
            day_work += 1
        elif dval == DN_REST:
            rest += 1
        elif dval == DN_SICK:
            sick += 1
            absent_list.append((name, dval))
        elif dval == DN_ROTATION:
            rotation += 1
            absent_list.append((name, dval))
        elif dval == DN_ABSENT:
            absent += 1
            absent_list.append((name, dval))
        elif dval == DN_MIGR:
            migr += 1
            absent_list.append((name, dval))

        if nval == DN_NIGHT:
            night_work += 1

    return {
        "day": day_work, "night": night_work, "rest": rest,
        "sick": sick, "rotation": rotation, "absent": absent, "migr": migr,
        "absent_list": absent_list, "total": len(active),
    }


def fire_employee(name: str, fire_date_str: str) -> bool:
    """
    Помечает сотрудника уволенным в листе «Сотрудники»:
    статус → 'уволен', дата увольнения → строка (ДД.ММ.ГГГГ или ДД.ММ).
    Строки в листах месяцев не трогает (история сохраняется).
    """
    try:
        ws = _open().worksheet(EMP_SHEET)
    except Exception:
        return False
    grid = ws.get_all_values()
    for i, r in enumerate(grid):
        if i == 0:
            continue  # шапка
        if len(r) >= 2 and r[1].strip() == name.strip():
            row = i + 1
            ws.update_cell(row, 3, EMP_STATUS_FIRED)
            ws.update_cell(row, 4, fire_date_str)
            _status_cache["data"] = None
            return True
    return False


def _fire_date_with_year(name: str, ref_year: int) -> str:
    """
    Берёт дату увольнения из листа «Сотрудники» и дополняет годом, если его нет.
    Правило края года: если месяц даты больше текущего — это прошлый год.
    Возвращает 'ДД.ММ.ГГГГ' или '' если даты нет.
    """
    for e in get_status_list():
        if e["name"].strip() == name.strip():
            raw = (e.get("fired_date") or "").strip()
            if not raw:
                return ""
            parts = raw.split(".")
            if len(parts) >= 3:
                return raw  # уже с годом
            if len(parts) == 2:
                try:
                    d, m = int(parts[0]), int(parts[1])
                except ValueError:
                    return raw
                now = datetime.now()
                y = ref_year
                if m > now.month:
                    y = ref_year - 1  # декабрь при вводе в январе
                return f"{d:02d}.{m:02d}.{y}"
            return raw
    return ""


def build_work_report(name: str, out_path: str, year: int = 2026) -> str | None:
    """
    Excel-график работы уволенного за месяцы, где он работал (модель день/ночь).
    Колонки: Число | День | Ночь. Возвращает путь или None.
    """
    import calendar as _cal
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    sp = _open()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Дата увольнения из листа «Сотрудники» (с автоподстановкой года)
    fire_date_full = _fire_date_with_year(name, year)

    thin = Border(*[Side(style="thin")] * 4)
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    any_data = False
    for month_idx, month_name in enumerate(MONTHS_RU, 1):
        try:
            ws_src = sp.worksheet(month_name)
        except Exception:
            continue
        grid = ws_src.get_all_values()
        emp_row = None
        for r in grid[FIRST_DATA_ROW - 1:]:
            if len(r) >= NAME_COL and r[NAME_COL - 1].strip() == name.strip():
                emp_row = r
                break
        if not emp_row:
            continue

        days = _cal.monthrange(year, month_idx)[1]
        # пары день/ночь начиная с FIRST_DAY_COL
        has_data = False
        rows_out = []
        for d in range(1, days + 1):
            d_col = FIRST_DAY_COL - 1 + (d - 1) * 2   # 0-based
            n_col = d_col + 1
            dval = emp_row[d_col].strip() if len(emp_row) > d_col else ""
            nval = emp_row[n_col].strip() if len(emp_row) > n_col else ""
            if dval or nval:
                has_data = True
            rows_out.append((d, dval, nval))
        if not has_data:
            continue

        any_data = True
        ws_out = wb.create_sheet(month_name)
        ws_out["A1"] = f"{name} — {month_name} {year}"
        ws_out["A1"].font = Font(bold=True, size=12)
        if fire_date_full:
            ws_out["A1"].value = f"{name} — {month_name} {year}   (уволен {fire_date_full})"
        for col, title in enumerate(["Число", "День", "Ночь"], 1):
            c = ws_out.cell(row=2, column=col, value=title)
            c.font = bold
            c.border = thin
            c.alignment = center
        for i, (d, dval, nval) in enumerate(rows_out, 1):
            ws_out.cell(row=i + 2, column=1, value=d).border = thin
            cd = ws_out.cell(row=i + 2, column=2, value=dval)
            cn = ws_out.cell(row=i + 2, column=3, value=nval)
            for c in (cd, cn):
                c.border = thin
                c.alignment = center
        ws_out.column_dimensions["A"].width = 7
        ws_out.column_dimensions["B"].width = 8
        ws_out.column_dimensions["C"].width = 8

    if not any_data:
        return None
    wb.save(out_path)
    return out_path


# ================= ПОЛЬЗОВАТЕЛИ (ДОСТУП) =================

_users_cache = {"data": None, "ts": 0}
_USERS_TTL = 20


def _ensure_users_sheet():
    """Создаёт лист «Пользователи», если его нет. Заполняет шапку."""
    sp = _open()
    try:
        sp.worksheet(USERS_SHEET)
        return
    except Exception:
        pass
    sp.batch_update({"requests": [
        {"addSheet": {"properties": {"title": USERS_SHEET}}}
    ]})
    ws = sp.worksheet(USERS_SHEET)
    ws.update("A1", [["chat_id", "Имя", "Роль"]])


def get_users(force: bool = False) -> list[dict]:
    """
    Список пользователей бота: [{chat_id, name, role}].
    При превышении квоты Sheets (429): если есть кэш — отдаём его (пусть
    устаревший, это лучше, чем ложное "нет доступа"); если кэша нет —
    поднимаем SheetsBusyError, чтобы вызывающий код показал понятную
    причину, а не решил, что пользователя нет в списке.
    """
    now = time.time()
    if (not force and _users_cache["data"] is not None
            and now - _users_cache["ts"] < _USERS_TTL):
        return _users_cache["data"]
    try:
        ws = _open().worksheet(USERS_SHEET)
        rows = ws.get_all_values()[1:]
    except Exception as e:
        if _is_quota_error(e):
            if _users_cache["data"] is not None:
                return _users_cache["data"]
            raise SheetsBusyError(
                "Google Sheets временно перегружен, попробуйте через минуту.") from e
        return []
    result = []
    for r in rows:
        if r and r[0].strip():
            try:
                cid = int(r[0].strip())
            except ValueError:
                continue
            result.append({
                "chat_id": cid,
                "name": (r[1].strip() if len(r) > 1 else ""),
                "role": (r[2].strip() if len(r) > 2 else ROLE_FOREMAN),
            })
    _users_cache["data"] = result
    _users_cache["ts"] = now
    return result


def is_allowed(chat_id: int) -> bool:
    """Разрешён ли пользователь (есть в списке)."""
    return any(u["chat_id"] == chat_id for u in get_users())


def get_role(chat_id: int) -> str | None:
    for u in get_users():
        if u["chat_id"] == chat_id:
            return u["role"]
    return None


def add_user(chat_id: int, name: str = "", role: str = ROLE_FOREMAN) -> bool:
    """Добавляет пользователя в лист «Пользователи»."""
    _ensure_users_sheet()
    if is_allowed(chat_id):
        return False  # уже есть
    ws = _open().worksheet(USERS_SHEET)
    ws.append_row([str(chat_id), name, role])
    _users_cache["data"] = None
    return True


def get_admins() -> list[int]:
    """
    chat_id всех админов.
    При перегрузке Sheets (SheetsBusyError) отдаём пустой список —
    вызывающий код (_send_access_request) просто не сможет уведомить
    админов в этот момент, но обработчик не падает необработанным 429.
    """
    try:
        users = get_users()
    except SheetsBusyError:
        return []
    return [u["chat_id"] for u in users if u["role"] == ROLE_ADMIN]


# ================= ПРОВЕРКИ И ОТЧЁТ =================

# Пороги
ABSENT_THRESHOLD = 2   # неявок за месяц (>=) → проблема
WEEKEND_THRESHOLD = 2  # выходных за месяц (>) → проблема


def _month_counts(date: datetime | None = None) -> dict:
    """
    Считает по каждому активному сотруднику коды за месяц (дневной слот).
    Один запрос на лист. Возвращает {ФИО: {код: count, ...}}.
    """
    date = date or datetime.now()
    active = set(get_employees(date))
    ws = _worksheet_for(date)
    grid = ws.get_all_values()
    days = calendar.monthrange(date.year, date.month)[1]

    result = {}
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) < NAME_COL:
            continue
        name = r[NAME_COL - 1].strip()
        if not name or name not in active:
            continue
        counts = {DN_DAY: 0, DN_NIGHT: 0, DN_REST: 0, DN_SICK: 0,
                  DN_ROTATION: 0, DN_ABSENT: 0, DN_MIGR: 0, DN_WEEKEND: 0}
        for d in range(days):
            d_idx = (FIRST_DAY_COL - 1) + d * 2
            n_idx = d_idx + 1
            dval = r[d_idx].strip() if len(r) > d_idx else ""
            nval = r[n_idx].strip() if len(r) > n_idx else ""
            if dval in counts:
                counts[dval] += 1
            if nval == DN_NIGHT:
                counts[DN_NIGHT] += 1
        result[name] = counts
    return result


def _month_counts_halved(date: datetime | None = None) -> dict:
    """
    Как _month_counts, но с разбивкой на половины месяца.
    Возвращает {ФИО: {"h1": {код:n}, "h2": {код:n}, "m": {код:n}}}.
    h1 — дни 1..15, h2 — дни 16..конец, m — весь месяц.
    """
    date = date or datetime.now()
    active = set(get_employees(date))
    ws = _worksheet_for(date)
    grid = ws.get_all_values()
    days = calendar.monthrange(date.year, date.month)[1]

    def _blank():
        return {DN_DAY: 0, DN_NIGHT: 0, DN_REST: 0, DN_SICK: 0,
                DN_ROTATION: 0, DN_ABSENT: 0, DN_MIGR: 0, DN_WEEKEND: 0}

    result = {}
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) < NAME_COL:
            continue
        name = r[NAME_COL - 1].strip()
        if not name or name not in active:
            continue
        h1, h2 = _blank(), _blank()
        for d in range(days):
            d_idx = (FIRST_DAY_COL - 1) + d * 2
            n_idx = d_idx + 1
            dval = r[d_idx].strip() if len(r) > d_idx else ""
            nval = r[n_idx].strip() if len(r) > n_idx else ""
            bucket = h1 if d < 15 else h2   # дни 1..15 → h1, 16.. → h2
            if dval in bucket:
                bucket[dval] += 1
            if nval == DN_NIGHT:
                bucket[DN_NIGHT] += 1
        m = {k: h1[k] + h2[k] for k in h1}
        result[name] = {"h1": h1, "h2": h2, "m": m}
    return result


def check_problems(date: datetime | None = None) -> list[dict]:
    """
    Проблемные сотрудники за месяц:
      - неявок (Н) >= ABSENT_THRESHOLD
      - выходных (В) > WEEKEND_THRESHOLD
    Возвращает [{name, absent, weekend, reasons:[...]}].
    """
    date = date or datetime.now()
    counts = _month_counts(date)
    problems = []
    for name, c in counts.items():
        reasons = []
        if c[DN_ABSENT] >= ABSENT_THRESHOLD:
            reasons.append(f"неявок {c[DN_ABSENT]}")
        if c[DN_WEEKEND] > WEEKEND_THRESHOLD:
            reasons.append(f"выходных {c[DN_WEEKEND]}")
        if reasons:
            problems.append({
                "name": name,
                "absent": c[DN_ABSENT],
                "weekend": c[DN_WEEKEND],
                "reasons": reasons,
            })
    return problems


def _daily_codes(date: datetime | None = None) -> dict:
    """
    Посуточные коды для табеля. Для каждого активного сотрудника —
    список по дням месяца: значение ячейки табеля.
      день Д + ночь НЧ  → 'С' (сутки)
      только ночь (НЧ)  → 'НЧ'  (в т.ч. когда день = О)
      дневной код (Д/Б/МЖ/МУ/В/Н) → он сам
      пусто → '' (в файле станет прочерком)
    Возвращает {ФИО: {"num": табельный_№, "codes": [коды по дням]}}.
    """
    date = date or datetime.now()
    active = set(get_employees(date))
    ws = _worksheet_for(date)
    grid = ws.get_all_values()
    days = calendar.monthrange(date.year, date.month)[1]

    result = {}
    for r in grid[FIRST_DATA_ROW - 1:]:
        if len(r) < NAME_COL:
            continue
        name = r[NAME_COL - 1].strip()
        if not name or name not in active:
            continue
        num_raw = r[NUM_COL - 1].strip() if len(r) > NUM_COL - 1 else ""
        try:
            num = int(num_raw)
        except ValueError:
            num = num_raw  # на случай нечислового значения
        row_codes = []
        for d in range(days):
            d_idx = (FIRST_DAY_COL - 1) + d * 2
            n_idx = d_idx + 1
            dval = r[d_idx].strip() if len(r) > d_idx else ""
            nval = r[n_idx].strip() if len(r) > n_idx else ""
            if dval == DN_DAY and nval == DN_NIGHT:
                code = "С"                      # сутки
            elif nval == DN_NIGHT:
                code = DN_NIGHT                 # только ночь (день О или пусто)
            elif dval:
                code = dval                     # дневной код
            else:
                code = ""                       # пусто
            row_codes.append(code)
        result[name] = {"num": num, "codes": row_codes}
    return result


def build_month_summary(out_path: str, date: datetime | None = None) -> str | None:
    """
    Свод в виде формы Т-13: шапка (ИП, подразделение, период),
    табельный номер, посуточная сетка одним блоком на весь месяц.
    В ячейке: С (день+ночь, жирный), НЧ, дневной код, Н (красный жирный), прочерк.
    Возвращает путь или None.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    date = date or datetime.now()
    daily = _daily_codes(date)
    if not daily:
        return None

    month_name = MONTHS_RU[date.month - 1]
    days = calendar.monthrange(date.year, date.month)[1]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Табель"
    ws.sheet_view.showGridLines = True  # видимая сетка листа (для пустых ячеек вокруг таблицы)

    thin = Border(*[Side(style="thin", color="999999")] * 4)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # --- Шапка формы ---
    ws["A1"] = "ИП Буц Сергей Юрьевич"
    ws["A1"].font = title_font
    ws["A2"] = "Подразделение: Новатэк Белокаменка/ПСМ"
    ws["A3"] = f"Табель учёта рабочего времени за {month_name} {date.year}"
    ws["A3"].font = bold
    ws["A4"] = (f"Отчётный период: 01.{date.month:02d}.{date.year} — "
                f"{days:02d}.{date.month:02d}.{date.year}")

    red_bold = Font(bold=True, color="CC0000")

    # Один блок на весь месяц: Таб.№ | ФИО | дни 1..days
    start_row = 6
    hc = ws.cell(start_row, 1, "Таб.№")
    hc.font = header_font; hc.alignment = center; hc.border = thin; hc.fill = header_fill
    hc = ws.cell(start_row, 2, "ФИО")
    hc.font = header_font; hc.alignment = center; hc.border = thin; hc.fill = header_fill
    for k in range(days):
        c = ws.cell(start_row, 3 + k, k + 1)
        c.font = header_font; c.alignment = center; c.border = thin; c.fill = header_fill

    row = start_row + 1
    ordered_names = sorted(daily.keys(), key=lambda n: n.strip().lower())
    for name in ordered_names:
        entry = daily[name]
        num = entry["num"]
        codes = entry["codes"]
        ws.cell(row, 1, num).border = thin
        ws.cell(row, 1).alignment = center
        nc = ws.cell(row, 2, name)
        nc.border = thin; nc.alignment = left
        for k in range(days):
            code = codes[k] if k < len(codes) else ""
            val = code if code else "—"
            cc = ws.cell(row, 3 + k, val)
            cc.border = thin
            cc.alignment = center
            if code == DN_ABSENT:
                cc.font = red_bold          # неявка — красный жирный
            elif code == "С":
                cc.font = bold              # сутки — жирный
        row += 1

    # --- Ширины ---
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    for col in range(3, 3 + days):
        ws.column_dimensions[get_column_letter(col)].width = 4

    # --- Сырые day/night значения для двух итоговых таблиц ниже ---
    active = set(get_employees(date))
    ws_src = _worksheet_for(date)
    grid_src = ws_src.get_all_values()

    day_totals = {code: [0] * days for code in
                  (DN_DAY, DN_NIGHT, DN_SICK, DN_ROTATION, DN_ABSENT, DN_MIGR, DN_WEEKEND)}
    emp_stats = {}  # ФИО -> {"shifts": кол-во смен, "days": отработано дней}

    for r in grid_src[FIRST_DATA_ROW - 1:]:
        if len(r) < NAME_COL:
            continue
        name = r[NAME_COL - 1].strip()
        if not name or name not in active:
            continue
        shifts = 0
        worked_days = 0
        for d in range(days):
            d_idx = (FIRST_DAY_COL - 1) + d * 2
            n_idx = d_idx + 1
            dval = r[d_idx].strip() if len(r) > d_idx else ""
            nval = r[n_idx].strip() if len(r) > n_idx else ""
            if dval in day_totals:
                day_totals[dval][d] += 1
            if nval == DN_NIGHT:
                day_totals[DN_NIGHT][d] += 1
            if dval == DN_DAY:
                shifts += 1
            if nval == DN_NIGHT:
                shifts += 1
            if dval == DN_DAY or nval == DN_NIGHT:
                worked_days += 1
        emp_stats[name] = {"shifts": shifts, "days": worked_days}

    # --- ТАБЛИЦА 1: Итог за день ---
    row += 2
    ws.cell(row, 1, "Итог за день").font = title_font
    row += 1
    hc = ws.cell(row, 1, "Показатель")
    hc.font = header_font; hc.alignment = center; hc.border = thin; hc.fill = header_fill
    for k in range(days):
        c = ws.cell(row, 2 + k, k + 1)
        c.font = header_font; c.alignment = center; c.border = thin; c.fill = header_fill
    row += 1

    day_row_labels = [
        (DN_DAY, "☀️ День"),
        (DN_NIGHT, "🌙 Ночь"),
        (DN_SICK, "🤒 Больничный"),
        (DN_ROTATION, "✈️ Межвахта"),
        (DN_ABSENT, "❌ Неявка"),
        (DN_MIGR, "📋 Мигр.учёт"),
        (DN_WEEKEND, "🏖 Выходной"),
    ]
    for code, label in day_row_labels:
        ws.cell(row, 1, label).border = thin
        ws.cell(row, 1).alignment = left
        for k in range(days):
            c = ws.cell(row, 2 + k, day_totals[code][k])
            c.border = thin
            c.alignment = center
        row += 1

    # --- ТАБЛИЦА 2: Общий итог за месяц ---
    row += 2
    ws.cell(row, 1, "Общий итог за месяц").font = title_font
    row += 1
    hc = ws.cell(row, 1, "ФИО")
    hc.font = header_font; hc.alignment = left; hc.border = thin; hc.fill = header_fill
    hc = ws.cell(row, 2, "Кол-во смен")
    hc.font = header_font; hc.alignment = center; hc.border = thin; hc.fill = header_fill
    hc = ws.cell(row, 3, "Отработано дней")
    hc.font = header_font; hc.alignment = center; hc.border = thin; hc.fill = header_fill
    row += 1
    for name in ordered_names:
        st = emp_stats.get(name, {"shifts": 0, "days": 0})
        nc = ws.cell(row, 1, name)
        nc.border = thin; nc.alignment = left
        sc = ws.cell(row, 2, st["shifts"])
        sc.border = thin; sc.alignment = center
        dc = ws.cell(row, 3, st["days"])
        dc.border = thin; dc.alignment = center
        row += 1

    wb.save(out_path)
    return out_path


def add_employees_from_xlsx(file_path: str) -> dict:
    """
    Массовое добавление сотрудников из Excel.
    Первый столбец = ФИО, первая строка = шапка (пропускается).
    Дата приёма = сегодня (внутри add_employee).
    Возвращает {added, skipped, invalid, fuzzy, fired}.
      fuzzy  — Фамилия+Имя совпали с существующим, но полное ФИО другое —
               не добавлены, требуют ручной проверки
      fired  — точное совпадение ФИО со статусом "уволен" —
               не добавлены и не восстановлены (статус не трогаем)
    """
    import openpyxl

    added, skipped, invalid, fuzzy, fired = [], [], [], [], []
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return {"added": [], "skipped": [], "invalid": [], "fuzzy": [], "fired": [],
                "error": f"Не удалось открыть файл: {e}"}

    ws = wb.active
    status_list = get_status_list()

    # Предзагрузка состояния ОДИН раз на всю пачку — дальше инкрементируем
    # в памяти, без обращений к Sheets API на каждого сотрудника.
    from googleapiclient.discovery import build as _build
    sp = _open()
    service = _build("sheets", "v4", credentials=_credentials())
    last_rows = _fetch_last_rows(service)
    try:
        emp_rows = sp.worksheet(EMP_SHEET).get_all_values()
        next_num = len([r for r in emp_rows[1:] if r and r[0].strip()]) + 1
        emp_new_row = len(emp_rows) + 1
    except Exception:
        next_num = len(status_list) + 1
        emp_new_row = len(status_list) + 2  # +1 шапка, +1 новая строка

    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False   # пропускаем шапку
            continue
        if not row or not row[0]:
            continue
        name = " ".join(str(row[0]).split())
        # минимальная валидация: 2+ слова кириллицей
        parts = name.split()
        if len(parts) < 2 or not all(
                all(ch.isalpha() or ch == "-" for ch in p) for p in parts):
            invalid.append(name)
            continue

        nm_full = name.strip().lower()
        exact = next((e for e in status_list if e["name"].strip().lower() == nm_full), None)
        if exact:
            if exact["status"] == EMP_STATUS_FIRED:
                fired.append(name)
            else:
                skipped.append(name)
            continue

        fm = find_fuzzy_matches(name, status_list)
        if fm:
            fuzzy.append({"new": name, "existing": [m["name"] for m in fm]})
            continue

        ok = add_employee(name, _skip_exists_check=True,
                           _next_num=next_num, _emp_new_row=emp_new_row,
                           _last_rows=last_rows)
        if ok:
            added.append(name)
            status_list.append({"name": name, "status": EMP_STATUS_ACTIVE, "fired_date": ""})
            # инкремент локального состояния — без повторного чтения API
            next_num += 1
            emp_new_row += 1
            for m in MONTHS_RU:
                last_rows[m] = last_rows.get(m, FIRST_DATA_ROW - 1) + 1
        else:
            skipped.append(name)

    return {"added": added, "skipped": skipped, "invalid": invalid,
            "fuzzy": fuzzy, "fired": fired}
